#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""온습도계 사진 → 체감온도 기록 프로그램"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import base64
import os
import json
import re
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk, ImageOps

CLAUDE_MODEL = "claude-sonnet-4-6"

def parse_date_from_filename(filename: str) -> str:
    """파일명에서 날짜 파싱: '20260506 (1).jpg' → '2026-05-06'"""
    m = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ""

# ── 체감온도 계산 (KOSHA/기상청 공식) ─────────────────────────────────────
import math as _math

def heat_index(Ta: float, RH: float) -> float:
    """KOSHA·기상청 여름철 체감온도 공식 (Stull 습구온도 이용)
    Ta: 기온(°C), RH: 상대습도(%)
    """
    # 습구온도 Tw (Stull 2011)
    Tw = (Ta * _math.atan(0.151977 * (RH + 8.313659) ** 0.5)
          + _math.atan(Ta + RH)
          - _math.atan(RH - 1.67633)
          + 0.00391838 * RH ** 1.5 * _math.atan(0.023101 * RH)
          - 4.686035)
    # 체감온도
    hi = (-0.2442
          + 0.55399 * Tw
          + 0.45535 * Ta
          - 0.0022  * Tw ** 2
          + 0.00278 * Tw * Ta
          + 3.0)
    return round(hi, 1)


def heat_index_label(hi: float) -> str:
    """KOSHA 폭염 영향예보 기준 단계
    미만31: 해당없음 / 31~33: 관심 / 33~35: 주의 / 35~38: 경고 / 38이상: 위험
    """
    if hi >= 38:
        return "위험"
    if hi >= 35:
        return "경고"
    if hi >= 33:
        return "주의"
    if hi >= 31:
        return "관심"
    return "-"


# ── 데이터 클래스 ──────────────────────────────────────────────────────────
class Record:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.filename = Path(filepath).name
        self.temperature: float | None = None
        self.humidity: float | None = None
        self.time_str: str = ""
        self.author: str = ""
        self.location: str = ""
        self.remarks: str = ""
        self.status: str = "대기중"

    @property
    def feels_like(self) -> float | None:
        if self.temperature is not None and self.humidity is not None:
            return heat_index(self.temperature, self.humidity)
        return None


# ── 메인 앱 ───────────────────────────────────────────────────────────────
class App(tk.Tk):
    COL_HEADERS = ["파일명", "측정시각", "온도(°C)", "습도(%)", "체감온도(°C)", "단계", "작성자", "측정장소", "비고", "상태"]
    COL_KEYS    = ["filename", "time_str", "temperature", "humidity", "feels_like", "_level", "author", "location", "remarks", "status"]
    COL_WIDTHS  = [180, 90, 80, 80, 95, 70, 80, 110, 110, 70]

    def __init__(self):
        super().__init__()
        self.title("온습도 체감온도 기록기")
        self.geometry("1300x720")
        self.minsize(960, 500)
        self.records: list[Record] = []
        self._thumb_cache: dict[str, ImageTk.PhotoImage] = {}
        # API 키 유무 확인 (없어도 수동 입력 모드로 실행 가능)
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        self.has_api = bool(api_key)
        self.client = anthropic.Anthropic(api_key=api_key) if self.has_api else None
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.configure(bg="#f0f4f8")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        font=("맑은 고딕", 10),
                        rowheight=26,
                        background="#ffffff",
                        fieldbackground="#ffffff")
        style.configure("Treeview.Heading",
                        font=("맑은 고딕", 10, "bold"),
                        background="#1a4f7a",
                        foreground="white")
        style.map("Treeview",
                  background=[("selected", "#cfe0f5")],
                  foreground=[("selected", "#000000")])

        # 상단 툴바
        toolbar = tk.Frame(self, bg="#1a4f7a", pady=8)
        toolbar.pack(fill="x")
        tk.Label(toolbar, text="🌡  온습도 체감온도 기록기",
                 bg="#1a4f7a", fg="white",
                 font=("맑은 고딕", 13, "bold")).pack(side="left", padx=16)
        for txt, cmd in [("📂  사진 불러오기", self.load_images),
                         ("💾  엑셀로 저장",   self.save_excel),
                         ("🗑  초기화",        self.clear_all)]:
            tk.Button(toolbar, text=txt, command=cmd,
                      bg="#2d6a9f", fg="white", relief="flat",
                      font=("맑은 고딕", 10), padx=10, pady=4,
                      activebackground="#3a87c8", activeforeground="white",
                      cursor="hand2").pack(side="left", padx=4)

        # AI 추출 버튼: API 키 있으면 활성, 없으면 비활성
        ai_btn_cfg = dict(
            text="🤖  AI 자동 추출",
            font=("맑은 고딕", 10), padx=10, pady=4, relief="flat"
        )
        if self.has_api:
            self.ai_btn = tk.Button(toolbar, command=self.extract_all,
                                    bg="#2d6a9f", fg="white",
                                    activebackground="#3a87c8",
                                    activeforeground="white",
                                    cursor="hand2", **ai_btn_cfg)
        else:
            self.ai_btn = tk.Button(toolbar,
                                    command=self._no_api_msg,
                                    bg="#666666", fg="#cccccc",
                                    cursor="arrow", **ai_btn_cfg)
        self.ai_btn.pack(side="left", padx=4)

        # 기본값 바
        dbar = tk.Frame(self, bg="#dce9f5", pady=5, padx=10)
        dbar.pack(fill="x")
        tk.Label(dbar, text="기본값 설정 →", bg="#dce9f5",
                 font=("맑은 고딕", 10, "bold")).pack(side="left", padx=4)
        for lbl, attr in [("작성자:", "dv_author"), ("측정장소:", "dv_location")]:
            tk.Label(dbar, text=lbl, bg="#dce9f5",
                     font=("맑은 고딕", 10)).pack(side="left", padx=(10, 2))
            sv = tk.StringVar()
            setattr(self, attr, sv)
            tk.Entry(dbar, textvariable=sv, width=16,
                     font=("맑은 고딕", 10)).pack(side="left")
        tk.Button(dbar, text="전체 적용", command=self._apply_defaults,
                  bg="#2d6a9f", fg="white", font=("맑은 고딕", 9),
                  relief="flat", padx=8).pack(side="left", padx=10)
        tk.Label(dbar,
                 text="  ※ 행을 더블클릭하면 개별 편집 가능",
                 bg="#dce9f5", fg="#555",
                 font=("맑은 고딕", 9)).pack(side="left")

        # 중앙 분할
        pane = tk.PanedWindow(self, orient="horizontal",
                              bg="#b0c8e0", sashwidth=5, sashrelief="flat")
        pane.pack(fill="both", expand=True, padx=6, pady=(4, 0))

        # 왼쪽: 이미지 미리보기
        lf = tk.Frame(pane, bg="#e4edf7", width=220)
        pane.add(lf, minsize=140)
        tk.Label(lf, text="미리보기", bg="#1a4f7a", fg="white",
                 font=("맑은 고딕", 10, "bold")).pack(fill="x")
        self.preview_lbl = tk.Label(lf, bg="#e4edf7",
                                    text="사진을 선택하세요",
                                    font=("맑은 고딕", 9), fg="#888",
                                    anchor="center")
        self.preview_lbl.pack(fill="both", expand=True, padx=4, pady=4)

        # 오른쪽: 테이블
        rf = tk.Frame(pane, bg="#f0f4f8")
        pane.add(rf, minsize=700)

        self.tree = ttk.Treeview(rf,
                                 columns=self.COL_HEADERS,
                                 show="headings",
                                 selectmode="browse")
        for h, w in zip(self.COL_HEADERS, self.COL_WIDTHS):
            self.tree.heading(h, text=h)
            self.tree.column(h, width=w, anchor="center", minwidth=50)

        vsb = ttk.Scrollbar(rf, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>",         self._on_double_click)

        self.tree.tag_configure("waiting",    foreground="#888888")
        self.tree.tag_configure("extracting", foreground="#1a66dd", font=("맑은 고딕", 10, "italic"))
        self.tree.tag_configure("done",       foreground="#1a7a3a")
        self.tree.tag_configure("error",      foreground="#cc2222")

        # 상태바
        self.status_sv = tk.StringVar(value="사진을 불러오세요.")
        tk.Label(self, textvariable=self.status_sv,
                 bg="#c0d4e8", anchor="w",
                 font=("맑은 고딕", 9), padx=10).pack(fill="x", side="bottom")

    def _no_api_msg(self):
        messagebox.showinfo(
            "API 키 필요",
            "AI 자동 추출을 사용하려면 ANTHROPIC_API_KEY 환경변수가 필요합니다.\n\n"
            "터미널에서 아래 명령어로 키를 설정한 후 프로그램을 다시 실행하세요:\n\n"
            "$env:ANTHROPIC_API_KEY = \"sk-ant-...\"\n"
            "python main.py\n\n"
            "※ 키 없이도 더블클릭으로 값을 직접 입력할 수 있습니다.")

    # ── 사진 불러오기 ─────────────────────────────────────────────────────
    def load_images(self):
        paths = filedialog.askopenfilenames(
            title="온습도계 사진 선택",
            filetypes=[("이미지 파일", "*.jpg *.jpeg *.png *.bmp"),
                       ("모든 파일", "*.*")])
        existing = {r.filepath for r in self.records}
        added = 0
        for p in paths:
            if p not in existing:
                r = Record(p)
                r.author   = self.dv_author.get()
                r.location = self.dv_location.get()
                self.records.append(r)
                added += 1
        self._refresh_table()
        self.status_sv.set(f"{added}개 추가됨 · 총 {len(self.records)}개")

    # ── 기본값 전체 적용 ──────────────────────────────────────────────────
    def _apply_defaults(self):
        a, l = self.dv_author.get(), self.dv_location.get()
        for r in self.records:
            if a: r.author   = a
            if l: r.location = l
        self._refresh_table()

    # ── AI 자동 추출 ──────────────────────────────────────────────────────
    def extract_all(self):
        pending = [r for r in self.records if r.status in ("대기중", "오류")]
        if not pending:
            messagebox.showinfo("알림", "추출할 항목이 없습니다.\n(완료된 항목은 건너뜁니다)")
            return
        self.status_sv.set(f"AI 추출 시작... ({len(pending)}개)")
        threading.Thread(target=self._extract_thread,
                         args=(pending,), daemon=True).start()

    def _extract_thread(self, pending: list[Record]):
        for i, rec in enumerate(pending):
            rec.status = "추출중"
            self.after(0, self._refresh_table)
            try:
                self._extract_one(rec)
            except Exception as e:
                import traceback
                rec.status = "오류"
                err_msg = str(e)
                # 로그 파일에 전체 오류 기록
                with open(r"C:\dev\온습도계기록\error.log", "a", encoding="utf-8") as lf:
                    lf.write(f"\n=== {rec.filename} ===\n")
                    lf.write(traceback.format_exc())
                self.after(0, lambda msg=err_msg, fn=rec.filename:
                           self.status_sv.set(f"오류 [{fn}]: {msg}"))
            self.after(0, lambda n=i+1, t=len(pending):
                       self.status_sv.set(f"AI 추출 중... {n}/{t}"))
        errors = [r for r in pending if r.status == "오류"]
        self.after(0, self._refresh_table)
        if errors:
            err_names = "\n".join(f"• {r.filename}" for r in errors)
            self.after(0, lambda:
                       messagebox.showerror("추출 실패",
                           f"다음 파일에서 오류가 발생했습니다:\n{err_names}\n\n"
                           f"하단 상태바에서 오류 내용을 확인하세요."))
        else:
            self.after(0, lambda:
                       self.status_sv.set(f"추출 완료 · 총 {len(self.records)}개"))

    def _extract_one(self, rec: Record):
        # EXIF 회전 보정 후 JPEG 바이트로 변환 (회전된 사진 대응)
        import io
        img = Image.open(rec.filepath)
        img = ImageOps.exif_transpose(img)
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=92)
        b64 = base64.standard_b64encode(buf.getvalue()).decode()

        prompt = (
            "이것은 CAS 브랜드 디지털 온습도계 사진입니다.\n"
            "LCD 디스플레이에서 다음 3가지 값을 읽어주세요:\n"
            "1. temperature: 위쪽 큰 숫자 → 온도(°C), 소수점 포함 (예: 25.1)\n"
            "2. humidity: 아래 오른쪽 큰 숫자 → 습도(%) 정수 (예: 87)\n"
            "3. time: 아래 왼쪽 작은 숫자 → 시각, AM/PM 변환해서 24시간제로 (예: '07:36'). 없으면 null\n"
            "사진이 회전되어 있어도 올바르게 읽어야 합니다.\n"
            "JSON만 출력 (다른 텍스트 없이):\n"
            "{\"temperature\": 25.1, \"humidity\": 87, \"time\": \"07:36\"}"
        )

        resp = self.client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=200,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image",
                     "source": {"type": "base64",
                                "media_type": "image/jpeg",
                                "data": b64}},
                    {"type": "text", "text": prompt}
                ]
            }]
        )

        text = resp.content[0].text.strip()
        m = re.search(r'\{.*?\}', text, re.DOTALL)
        if not m:
            raise ValueError(f"JSON 파싱 실패: {text[:80]}")
        data = json.loads(m.group())

        rec.temperature = float(data["temperature"]) if data.get("temperature") is not None else None
        rec.humidity    = float(data["humidity"])    if data.get("humidity") is not None else None

        # 날짜는 파일명에서, 시각은 AI에서
        date_str = parse_date_from_filename(rec.filename)
        t = data.get("time")
        time_part = t if (t and t != "null") else datetime.now().strftime("%H:%M")
        rec.time_str = f"{date_str} {time_part}".strip() if date_str else time_part

        if not rec.author:
            rec.author   = self.dv_author.get()
        if not rec.location:
            rec.location = self.dv_location.get()
        rec.status = "완료"

    # ── 테이블 갱신 ───────────────────────────────────────────────────────
    def _refresh_table(self):
        sel = self.tree.selection()
        sel_idx = int(self.tree.item(sel[0], "text")) if sel else -1

        self.tree.delete(*self.tree.get_children())
        for i, r in enumerate(self.records):
            fl = r.feels_like
            vals = [
                r.filename,
                r.time_str,
                f"{r.temperature:.1f}" if r.temperature is not None else "",
                f"{r.humidity:.0f}"    if r.humidity    is not None else "",
                f"{fl:.1f}"            if fl            is not None else "",
                heat_index_label(fl)   if fl            is not None else "",
                r.author,
                r.location,
                r.remarks,
                r.status,
            ]
            tag = {"대기중": "waiting", "추출중": "extracting",
                   "완료": "done", "오류": "error"}.get(r.status, "")
            iid = self.tree.insert("", "end", text=str(i), values=vals, tags=(tag,))
            if i == sel_idx:
                self.tree.selection_set(iid)
                self.tree.see(iid)

    # ── 미리보기 ──────────────────────────────────────────────────────────
    def _on_select(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(self.tree.item(sel[0], "text"))
        fp  = self.records[idx].filepath
        if fp not in self._thumb_cache:
            try:
                img = Image.open(fp)
                img = ImageOps.exif_transpose(img)
                img.thumbnail((210, 340))
                self._thumb_cache[fp] = ImageTk.PhotoImage(img)
            except Exception:
                self._thumb_cache[fp] = None
        ph = self._thumb_cache.get(fp)
        if ph:
            self.preview_lbl.config(image=ph, text="")
            self.preview_lbl.image = ph
        else:
            self.preview_lbl.config(image="", text="미리보기 불가")

    # ── 더블클릭 편집 ─────────────────────────────────────────────────────
    def _on_double_click(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(self.tree.item(sel[0], "text"))
        self._edit_dialog(self.records[idx])

    def _edit_dialog(self, rec: Record):
        dlg = tk.Toplevel(self)
        dlg.title(f"편집  –  {rec.filename}")
        dlg.geometry("400x330")
        dlg.resizable(False, False)
        dlg.grab_set()

        fields = [
            ("측정시각",  "time_str",    str),
            ("온도 (°C)", "temperature", float),
            ("습도 (%)",  "humidity",    float),
            ("작성자",    "author",      str),
            ("측정장소",  "location",    str),
            ("비고사항",  "remarks",     str),
        ]
        svars: dict[str, tk.StringVar] = {}
        for ri, (lbl, key, _) in enumerate(fields):
            tk.Label(dlg, text=lbl, anchor="e", width=12,
                     font=("맑은 고딕", 10)).grid(row=ri, column=0, padx=10, pady=7, sticky="e")
            val = getattr(rec, key)
            sv  = tk.StringVar(value="" if val is None else str(val))
            svars[key] = sv
            tk.Entry(dlg, textvariable=sv, width=26,
                     font=("맑은 고딕", 10)).grid(row=ri, column=1, padx=10, pady=7)

        def _save():
            try:
                rec.time_str = svars["time_str"].get()
                t = svars["temperature"].get().strip()
                rec.temperature = float(t) if t else None
                h = svars["humidity"].get().strip()
                rec.humidity = float(h) if h else None
                rec.author   = svars["author"].get()
                rec.location = svars["location"].get()
                rec.remarks  = svars["remarks"].get()
                if rec.temperature is not None and rec.humidity is not None:
                    rec.status = "완료"
                self._refresh_table()
                dlg.destroy()
            except ValueError:
                messagebox.showerror("오류", "온도/습도는 숫자로 입력하세요.", parent=dlg)

        n = len(fields)
        tk.Button(dlg, text="저장", command=_save,
                  bg="#2d6a9f", fg="white",
                  font=("맑은 고딕", 10), padx=18,
                  relief="flat").grid(row=n, column=0, pady=14, sticky="e")
        tk.Button(dlg, text="취소", command=dlg.destroy,
                  font=("맑은 고딕", 10), padx=12).grid(row=n, column=1, pady=14, sticky="w")

    # ── 엑셀 저장 ─────────────────────────────────────────────────────────
    def save_excel(self):
        done = [r for r in self.records if r.status == "완료"]
        if not done:
            messagebox.showwarning("알림", "저장할 완료된 데이터가 없습니다.")
            return

        default_name = f"온습도기록_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="엑셀 파일 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=default_name)
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "온습도 기록"

        thin    = Side(style="thin", color="AAAAAA")
        border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr_fill = PatternFill("solid", fgColor="1A4F7A")
        hdr_font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
        data_font = Font(name="맑은 고딕", size=10)

        # 제목 행
        ws.merge_cells("A1:H1")
        ws["A1"] = f"체감온도 기록 – {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}"
        ws["A1"].font      = Font(bold=True, size=13, name="맑은 고딕")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 30

        # 헤더 행
        headers = ["작성자", "측정시각", "측정장소",
                   "온도(°C)", "습도(%)", "체감온도(°C)", "단계", "비고사항"]
        col_w   = [14, 12, 18, 11, 10, 14, 10, 24]
        for c, (h, w) in enumerate(zip(headers, col_w), 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[2].height = 22

        # 데이터 행
        # KOSHA 폭염 영향예보 공식 색상
        level_colors = {
            "위험": ("ED1C24", "FFFFFF"),  # 빨강
            "경고": ("F7941D", "FFFFFF"),  # 주황
            "주의": ("FFD600", "000000"),  # 노랑
            "관심": ("00A651", "FFFFFF"),  # 초록
        }
        for ri, rec in enumerate(done, 3):
            fl  = rec.feels_like
            lev = heat_index_label(fl) if fl is not None else ""
            row_data = [rec.author, rec.time_str, rec.location,
                        rec.temperature, rec.humidity, fl, lev, rec.remarks]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=c, value=val)
                cell.alignment = center
                cell.border    = border
                cell.font      = data_font
                if c in (4, 5, 6) and val is not None:
                    cell.number_format = "0.0"
            # 단계 열(G=7) 색상
            lev_cell = ws.cell(row=ri, column=7)
            if lev in level_colors:
                bg, fg = level_colors[lev]
                lev_cell.fill = PatternFill("solid", fgColor=bg)
                lev_cell.font = Font(color=fg, bold=True,
                                     name="맑은 고딕", size=10)
            ws.row_dimensions[ri].height = 20

        # ── 사진대지 (columns I=9 ~ L=12) ────────────────────────────────────
        import io as _io

        PHOTO_BASE  = 9    # 첫 번째 사진 열 (I)
        PHOTO_CH_W  = 28   # 열 너비 (chars ≈ 200px)
        PHOTO_PX_W  = 196  # 사진 가로 최대 px
        PHOTO_PX_H  = 260  # 사진 세로 최대 px

        # 열 너비 설정
        for i in range(4):
            ws.column_dimensions[get_column_letter(PHOTO_BASE + i)].width = PHOTO_CH_W

        # 사진대지 헤더 (I1:L1)
        ph_end = get_column_letter(PHOTO_BASE + 3)
        ws.merge_cells(f"I1:{ph_end}1")
        ph_hdr = ws["I1"]
        ph_hdr.value     = "사진대지"
        ph_hdr.fill      = hdr_fill
        ph_hdr.font      = hdr_font
        ph_hdr.alignment = center

        # 번호 레이블 (I2:L2) – 1, 2, 3, 4
        num_fill = PatternFill("solid", fgColor="DCE9F5")
        for i in range(4):
            c = ws.cell(row=2, column=PHOTO_BASE + i, value=str(i + 1))
            c.fill      = num_fill
            c.font      = Font(bold=True, name="맑은 고딕", size=11)
            c.alignment = center
            c.border    = border

        # 파일명 (N) → 사진 슬롯 매핑
        photo_map: dict[int, Record] = {}
        for rec in done:
            m = re.search(r'\((\d+)\)', rec.filename)
            if m:
                n = int(m.group(1))
                if 1 <= n <= 4:
                    photo_map[n] = rec

        if photo_map:
            # 사진은 데이터 마지막 행 + 2 행부터 시작
            photo_row = 3 + len(done) + 1

            # 사진대지 번호 레이블 행
            for i in range(4):
                c = ws.cell(row=photo_row, column=PHOTO_BASE + i, value=str(i + 1))
                c.fill      = num_fill
                c.font      = Font(bold=True, name="맑은 고딕", size=11)
                c.alignment = center
                c.border    = border
            ws.row_dimensions[photo_row].height = 18

            img_row = photo_row + 1
            ws.row_dimensions[img_row].height = PHOTO_PX_H * 0.75  # px → pt

            for n, rec in photo_map.items():
                col = PHOTO_BASE + n - 1
                try:
                    pil = Image.open(rec.filepath)
                    pil = ImageOps.exif_transpose(pil)
                    pil.thumbnail((PHOTO_PX_W, PHOTO_PX_H))
                    buf = _io.BytesIO()
                    pil.convert("RGB").save(buf, format="PNG")
                    buf.seek(0)
                    xi        = XLImage(buf)
                    xi.width  = pil.width
                    xi.height = pil.height
                    ws.add_image(xi, f"{get_column_letter(col)}{img_row}")
                except Exception:
                    ws.cell(row=img_row, column=col, value="이미지 오류")

        ws.freeze_panes = "A3"
        wb.save(path)
        messagebox.showinfo("저장 완료",
                            f"{len(done)}개 기록이 저장되었습니다.\n\n{path}")
        self.status_sv.set(f"엑셀 저장 완료 · {path}")

    # ── 초기화 ────────────────────────────────────────────────────────────
    def clear_all(self):
        if not self.records:
            return
        if messagebox.askyesno("초기화 확인", "모든 기록을 지우시겠습니까?"):
            self.records.clear()
            self._thumb_cache.clear()
            self._refresh_table()
            self.preview_lbl.config(image="", text="사진을 선택하세요")
            self.status_sv.set("초기화되었습니다.")


# ── 진입점 ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except anthropic.AuthenticationError:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        tk.messagebox.showerror(
            "API 키 오류",
            "ANTHROPIC_API_KEY 환경변수가 설정되지 않았습니다.\n\n"
            "터미널에서 다음을 실행하세요:\n"
            "  $env:ANTHROPIC_API_KEY = 'your-key-here'\n"
            "  python main.py")
        root.destroy()
