#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""체감온도 기록관리 대장 – Streamlit"""

import streamlit as st
from google import genai as google_genai
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageOps
import pandas as pd
import io, re, json, math, os, calendar, time, base64
import plotly.graph_objects as go
from datetime import datetime, date, timedelta
from collections import defaultdict

DAYS_KO = ["월", "화", "수", "목", "금", "토", "일"]
SLOTS   = ["오전1", "오전2", "오후1", "오후2"]
ACTIONS = ["N/A", "추가휴식시간부여", "보냉장구지급", "작업시간대조정", "작업중지", "기타"]
HISTORY_FILE = os.path.join(os.path.dirname(__file__), ".app_history.json")

st.set_page_config(page_title="체감온도 기록관리 대장", page_icon="🌡",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""<style>
  .stApp { background:#f5f5f7; }
  section[data-testid="stSidebar"] > div { background:#ffffff !important; border-right:1px solid #e8e8ed; }

  .hero {
    background: linear-gradient(135deg, #1c1c1e 0%, #2c5282 100%);
    padding:22px 28px; border-radius:18px; color:white; margin-bottom:18px;
    box-shadow:0 4px 20px rgba(0,0,0,.18);
  }
  .hero h1 { margin:0; font-size:1.55rem; font-weight:700; letter-spacing:-.4px; }
  .hero-meta { margin:10px 0 0; display:flex; flex-wrap:wrap; gap:8px; }
  .chip { background:rgba(255,255,255,.15); backdrop-filter:blur(4px);
          border:1px solid rgba(255,255,255,.25);
          padding:3px 12px; border-radius:20px; font-size:.8rem; }

  .mac-card { background:#fff; border-radius:14px; padding:18px 22px;
               box-shadow:0 1px 8px rgba(0,0,0,.07); border:1px solid #e8e8ed;
               margin-bottom:14px; }

  .kpi-card { background:white; border:1px solid #e8e8ed; border-radius:13px;
               padding:14px 16px; box-shadow:0 1px 5px rgba(0,0,0,.05); text-align:center; }
  .kpi-val  { font-size:1.6rem; font-weight:700; color:#1c1c1e; line-height:1.1; }
  .kpi-lbl  { font-size:.72rem; color:#86868b; margin-top:3px; }
  .kpi-card-soft { background:linear-gradient(135deg,#e8f4fd 0%,#dbeafe 100%);
                   border:1px solid #bfdbfe; border-radius:13px; padding:14px 16px;
                   box-shadow:0 1px 5px rgba(0,0,0,.05); text-align:center; }
  .kpi-card-soft .kpi-val { color:#1e40af; }
  .kpi-card-soft .kpi-lbl { color:#6b7280; margin-top:3px; }
  .kpi-card-hl { background:linear-gradient(135deg,#FFAB76 0%,#FF7043 100%);
                 border:none; border-radius:13px; padding:14px 16px;
                 box-shadow:0 4px 14px rgba(255,112,67,.25); text-align:center; }
  .kpi-card-hl .kpi-val { color:white; }
  .kpi-card-hl .kpi-lbl { color:rgba(255,255,255,.85); }

  .week-tag { display:inline-block; background:#0071e3; color:#fff;
               border-radius:8px; padding:3px 12px; font-size:.82rem; font-weight:600; }

  div[data-testid="stFileUploader"] {
    border:1.5px dashed #c7c7cc; border-radius:13px; padding:14px; background:#fafafa; }

  .stButton>button { border-radius:10px !important; font-weight:600 !important; font-size:.87rem !important; }
  .stDataEditor   { border-radius:12px !important; }

  .kosha-row { display:flex; align-items:center; gap:8px; padding:4px 0; font-size:.83rem; }
  .dot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
</style>""", unsafe_allow_html=True)

# ── 헬퍼 함수 ──────────────────────────────────────────────────────────────
def heat_index(Ta: float, RH: float) -> float:
    Tw = (Ta * math.atan(0.151977*(RH+8.313659)**0.5)
          + math.atan(Ta+RH) - math.atan(RH-1.67633)
          + 0.00391838*RH**1.5*math.atan(0.023101*RH) - 4.686035)
    return round(-0.2442+0.55399*Tw+0.45535*Ta-0.0022*Tw**2+0.00278*Tw*Ta+3.0, 1)

LEVEL_BADGE = {"위험":"🔴 위험","경고":"🟠 경고","주의":"🟡 주의","관심":"🟢 관심","-":"-","":""}

def heat_label(hi: float) -> str:
    if hi >= 38: return "위험"
    if hi >= 35: return "경고"
    if hi >= 33: return "주의"
    if hi >= 31: return "관심"
    return "-"

def auto_action(hi) -> str:
    return "N/A" if (hi is None or hi < 31) else "추가휴식시간부여"

def parse_date(filename: str) -> str:
    # YYYYMMDD (20xx년도, 앞뒤 숫자 없어야 함)
    m = re.search(r'(?<!\d)(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)', filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # YYYY-MM-DD 또는 YYYY.MM.DD
    m = re.search(r'(20\d{2})[-./](0[1-9]|1[0-2])[-./](0[1-9]|[12]\d|3[01])', filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # YYMMDD (6자리, 20xx 세기로 해석)
    m = re.search(r'(?<!\d)(\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)', filename)
    if m:
        return f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ""

def resize_image(raw: bytes, max_px: int = 1568) -> bytes:
    img = Image.open(io.BytesIO(raw))
    img = ImageOps.exif_transpose(img)
    w, h = img.size
    if max(w, h) > max_px:
        s = max_px / max(w, h)
        img = img.resize((int(w*s), int(h*s)), Image.LANCZOS)
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=85)
    return buf.getvalue()

def get_month_weeks(year: int, month: int):
    first = date(year, month, 1)
    last  = date(year, month, calendar.monthrange(year, month)[1])
    mon   = first - timedelta(days=first.weekday())
    weeks, n = [], 1
    while mon <= last:
        sun = mon + timedelta(days=6)
        if sun >= first:
            # 월 경계를 자르지 않고 월~일 전체를 한 주로 표시
            lbl = (f"{n}째주  "
                   f"{mon.month}/{mon.day}({DAYS_KO[mon.weekday()]}) "
                   f"~ {sun.month}/{sun.day}({DAYS_KO[sun.weekday()]})")
            weeks.append((n, lbl, mon))
        mon += timedelta(days=7); n += 1
    return weeks

def week_label_ko(n: int) -> str:
    return ["첫째", "둘째", "셋째", "넷째", "다섯째"][min(n-1, 4)] + "주"

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"현장명": [], "업체명": [], "위치": [], "현장코드_map": {}, "records": []}

def save_history(meta: dict, records: list, monday: date = None):
    h = load_history()
    for k in ["현장명", "업체명", "위치"]:
        v = meta.get(k, "").strip()
        if v and v not in h[k]:
            h[k].append(v)
    code = meta.get("현장코드", "").strip()
    name = meta.get("현장명", "").strip()
    if code and name:
        if "현장코드_map" not in h:
            h["현장코드_map"] = {}
        h["현장코드_map"][code] = name
    mon_str = monday.isoformat() if monday else ""
    for rec in records:
        if not rec.get("_done"):
            continue
        slim = {k: v for k, v in rec.items() if k not in ("_bytes", "_bytes_b64")}
        if rec.get("_bytes"):
            slim["_bytes_b64"] = base64.b64encode(rec["_bytes"]).decode("ascii")
        slim["_위치"]     = meta.get("위치", "")
        slim["_현장명"]   = meta.get("현장명", "")
        slim["_업체명"]   = meta.get("업체명", "")
        slim["_현장코드"] = meta.get("현장코드", "")
        slim["_monday"]   = mon_str
        idx = next(
            (i for i, r in enumerate(h["records"])
             if r.get("_date") == slim.get("_date")
             and r.get("_slot") == slim.get("_slot")
             and r.get("_위치","") == slim.get("_위치","")),
            None,
        )
        if idx is not None:
            h["records"][idx] = slim
        else:
            h["records"].append(slim)
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(h, f, ensure_ascii=False, default=str)

# ── Excel 생성 ────────────────────────────────────────────────────────────
def build_excel(records: list, meta: dict, monday: date) -> bytes:
    from openpyxl.worksheet.page import PageMargins
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "체감온도 기록관리 대장"

    # ── 스타일 ──────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="B8CCE0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=False)

    title_fill = PatternFill("solid", fgColor="1E3A5F")
    title_font = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=13)
    hfill = PatternFill("solid", fgColor="2D6A9F")
    hfont = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=9)
    pfill = PatternFill("solid", fgColor="1E6091")
    pfont = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=9)
    mfill = PatternFill("solid", fgColor="EBF3FA")
    mfont = Font(bold=True, name="맑은 고딕", size=9, color="1E3A5F")
    dfont = Font(name="맑은 고딕", size=9, color="1C2B3A")
    date_font = Font(bold=True, name="맑은 고딕", size=9, color="1E3A5F")
    dfill0 = PatternFill("solid", fgColor="FFFFFF")
    dfill1 = PatternFill("solid", fgColor="F0F7FB")
    lv_colors = {"위험":("FF3B30","FFFFFF"), "경고":("FF9500","FFFFFF"),
                 "주의":("FFD60A","1C2B3A"), "관심":("30D158","FFFFFF")}

    ROW_H = 28
    NCOL  = 11
    PC    = NCOL + 1
    PCW   = 21
    PPH   = int(ROW_H * 4 * 4 / 3) - 4    # ≈145px
    PPW   = int(PCW * 7) - 4               # ≈143px

    HDR     = ["작성일","구분","측정시각","온도(°C)","습도(%)","체감온도(°C)",
               "단계","조치사항","기타내용","측정자","비고"]
    CWIDTHS = [10, 7, 9, 7.5, 7.5, 9.5, 7, 14, 12, 9, 14]

    for i, w in enumerate(CWIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(4):
        ws.column_dimensions[get_column_letter(PC+i)].width = PCW

    # ── 주차 정보 ─────────────────────────────────────────────────────────
    wk      = get_month_weeks(monday.year, monday.month)
    wk_info = next((w for w in wk if w[2] == monday), None)
    wk_lbl  = (f"{monday.year}년 {monday.month}월 {week_label_ko(wk_info[0])}"
               if wk_info else "")

    # ── 행 1 : 제목 ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(NCOL)}1")
    c = ws["A1"]; c.value = "체감온도 기록관리 대장"
    c.font = title_font; c.alignment = ctr; c.fill = title_fill; c.border = bdr
    ws.merge_cells(f"{get_column_letter(PC)}1:{get_column_letter(PC+3)}2")
    ph = ws[f"{get_column_letter(PC)}1"]
    ph.value = "사진대지"; ph.fill = pfill; ph.font = pfont; ph.alignment = ctr; ph.border = bdr
    ws.row_dimensions[1].height = 22

    # ── 행 2 : 메타 정보 ──────────────────────────────────────────────────
    ws.merge_cells("A2:C2"); ws.merge_cells("D2:F2")
    ws.merge_cells("G2:I2"); ws.merge_cells("J2:K2")
    meta_align = Alignment(horizontal="left", vertical="center",
                           wrap_text=False, shrink_to_fit=True)
    for addr, val in [("A2", f"현장명: {meta.get('현장명','')}"),
                      ("D2", f"업체명: {meta.get('업체명','')}"),
                      ("G2", f"측정위치: {meta.get('위치','')}"),
                      ("J2", wk_lbl)]:
        c = ws[addr]; c.value = val
        c.font = mfont; c.alignment = meta_align; c.border = bdr; c.fill = mfill
    ws.row_dimensions[2].height = 18

    # ── 행 3 : 컬럼 헤더 ──────────────────────────────────────────────────
    for ci, h in enumerate(HDR, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = ctr; c.border = bdr
    for i, lbl in enumerate(SLOTS):
        c = ws.cell(row=3, column=PC+i, value=lbl)
        c.fill = pfill; c.font = pfont; c.alignment = ctr; c.border = bdr
    ws.row_dimensions[3].height = 16

    # ── 행 4~ : 7일 × 4슬롯 데이터 ──────────────────────────────────────
    rec_map: dict[tuple, dict] = {}
    for r in records:
        rec_map[(r.get("_date",""), r.get("_slot",""))] = r

    row_num = 4
    for day_i in range(7):
        d    = monday + timedelta(days=day_i)
        rs   = row_num
        re_  = row_num + 3
        dfill = dfill0 if day_i % 2 == 0 else dfill1

        for slot in SLOTS:
            rec = rec_map.get((d.isoformat(), slot), {})
            fl  = rec.get("체감온도(°C)")
            lv  = rec.get("단계","")
            act = rec.get("조치사항","")
            row_data = [
                f"{d.month}/{d.day}({DAYS_KO[d.weekday()]})",
                slot,
                rec.get("측정시각",""),
                rec.get("온도(°C)","") if rec.get("온도(°C)") is not None else "",
                rec.get("습도(%)" ,"") if rec.get("습도(%)") is not None else "",
                fl if fl is not None else "",
                lv,
                act if act != "기타" else "기타",
                rec.get("기타내용","") if act == "기타" else "",
                rec.get("측정자",""),
                rec.get("비고",""),
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.font = dfont; c.alignment = ctr; c.border = bdr; c.fill = dfill
                if ci in (4,5,6) and val != "":
                    c.number_format = "0.0"
            if lv in lv_colors:
                bg, fg = lv_colors[lv]
                lc = ws.cell(row=row_num, column=7)
                lc.fill = PatternFill("solid", fgColor=bg)
                lc.font = Font(color=fg, bold=True, name="맑은 고딕", size=9)
            row_num += 1

        # 작성일 병합
        ws.merge_cells(start_row=rs, start_column=1, end_row=re_, end_column=1)
        dc = ws.cell(row=rs, column=1)
        dc.font = date_font; dc.alignment = ctr; dc.fill = dfill; dc.border = bdr

        # 사진대지: 슬롯별 4행 병합 후 사진 삽입
        for si, slot in enumerate(SLOTS):
            col   = PC + si
            col_l = get_column_letter(col)
            ws.merge_cells(start_row=rs, start_column=col, end_row=re_, end_column=col)
            tc = ws.cell(row=rs, column=col)
            tc.border = bdr; tc.alignment = ctr; tc.fill = dfill

            rec = rec_map.get((d.isoformat(), slot))
            if rec and rec.get("_bytes"):
                try:
                    pil = Image.open(io.BytesIO(rec["_bytes"])).convert("RGB")
                    pil = pil.resize((PPW, PPH), Image.LANCZOS)
                    buf = io.BytesIO(); pil.save(buf, format="PNG"); buf.seek(0)
                    xi = XLImage(buf); xi.width = PPW; xi.height = PPH
                    ws.add_image(xi, f"{col_l}{rs}")
                except Exception:
                    tc.value = "⚠"

    # 행 높이 통일
    for r in range(4, row_num):
        ws.row_dimensions[r].height = ROW_H

    # ── KOSHA 범례 ────────────────────────────────────────────────────────
    leg = row_num
    for lbl, c1, c2, bg, fg in [
        ("KOSHA 폭염 단계", 1, 2, "1E3A5F", "FFFFFF"),
        ("■ 관심  31°C~",   3, 4, "30D158", "FFFFFF"),
        ("■ 주의  33°C~",   5, 6, "FFD60A", "1C2B3A"),
        ("■ 경고  35°C~",   7, 8, "FF9500", "FFFFFF"),
        ("■ 위험  38°C~",   9, 11, "FF3B30", "FFFFFF"),
    ]:
        ws.merge_cells(start_row=leg, start_column=c1, end_row=leg, end_column=c2)
        c = ws.cell(row=leg, column=c1, value=lbl)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color=fg, bold=True, name="맑은 고딕", size=8)
        c.alignment = ctr; c.border = bdr
    ws.row_dimensions[leg].height = 14
    row_num += 1

    # ── 서명란 ──────────────────────────────────────────────────────────────
    sig = row_num
    for label, c1, c2 in [("작성자", 1, 3), ("검토자", 4, 7), ("승인자", 8, 11)]:
        ws.merge_cells(start_row=sig, start_column=c1, end_row=sig, end_column=c2)
        c = ws.cell(row=sig, column=c1, value=f"{label}:                              (인)")
        c.border = bdr
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        c.font = Font(name="맑은 고딕", size=9, color="1C2B3A")
    ws.row_dimensions[sig].height = ROW_H * 2
    row_num += 1

    # ── 인쇄 설정 ────────────────────────────────────────────────────────
    ws.print_area = f"A1:{get_column_letter(PC+3)}{row_num-1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                  header=0.2, footer=0.2)
    ws.print_title_rows = "1:3"
    ws.freeze_panes = "A4"
    out = io.BytesIO(); wb.save(out); return out.getvalue()

# ── AI 추출 함수 ──────────────────────────────────────────────────────────
PROMPT = (
    "이것은 CAS 브랜드 디지털 온습도계 사진입니다.\n"
    "LCD 디스플레이에서 다음 3가지 값을 읽어주세요:\n"
    "1. temperature: 위쪽 큰 숫자 → 온도(°C), 소수점 포함 (예: 25.1)\n"
    "2. humidity: 아래 오른쪽 큰 숫자 → 습도(%) 정수 (예: 87)\n"
    "3. time: 아래 왼쪽 작은 숫자 → 시각, AM/PM이 있으면 24시간제로 변환 "
    "(예: 7:36 AM → '07:36', 2:05 PM → '14:05'). 안 보이면 null\n"
    "사진이 회전되어 있어도 올바르게 읽을 것.\n"
    "JSON만 출력: {\"temperature\":25.1,\"humidity\":87,\"time\":\"07:36\"}"
)

def _extract_one(client, rec: dict):
    last_err = None
    for attempt in range(3):
        try:
            img = Image.open(io.BytesIO(rec["_bytes"])).convert("RGB")
            resp = client.models.generate_content(model="gemini-2.5-flash", contents=[PROMPT, img])
            break
        except Exception as e:
            last_err = e
            if attempt < 2:
                time.sleep(2 ** attempt)
    else:
        raise last_err

    text = resp.text.strip()
    m = re.search(r'\{.*?\}', text, re.DOTALL)
    if not m:
        raise ValueError(f"JSON 파싱 실패: {text[:60]}")
    d  = json.loads(m.group())
    T  = float(d["temperature"]) if d.get("temperature") is not None else None
    RH = float(d["humidity"])    if d.get("humidity")    is not None else None
    t  = d.get("time")

    rec["온도(°C)"] = T
    rec["습도(%)"]  = RH
    if T is not None and RH is not None:
        hi = heat_index(T, RH)
        rec["체감온도(°C)"] = hi
        rec["단계"]        = heat_label(hi)
        rec["조치사항"]    = auto_action(hi)

    date_p = rec.get("_date","")
    if t and t != "null":
        time_p = t
    else:
        try:
            raw_img = Image.open(io.BytesIO(rec["_bytes"]))
            exif = raw_img._getexif() or {}
            dt_str = exif.get(36867) or exif.get(306) or ""
            time_p = datetime.strptime(dt_str, "%Y:%m:%d %H:%M:%S").strftime("%H:%M") if dt_str else datetime.now().strftime("%H:%M")
        except Exception:
            time_p = datetime.now().strftime("%H:%M")
    rec["측정시각"] = time_p
    rec["_done"] = True

# ── 세션 상태 ──────────────────────────────────────────────────────────────
if "records" not in st.session_state:
    st.session_state.records = []

if "meta" not in st.session_state:
    now = datetime.now()
    weeks = get_month_weeks(now.year, now.month)
    today = date.today()
    cur_week = next((w for w in weeks if w[2] <= today <= w[2]+timedelta(days=6)), weeks[0] if weeks else None)
    st.session_state.meta = {
        "현장코드": "", "현장명": "", "업체명": "", "위치": "", "측정자": "",
        "year": now.year, "month": now.month,
        "week_n": cur_week[0] if cur_week else 1,
        "monday": cur_week[2] if cur_week else date(now.year, now.month, 1),
    }

if "file_history" not in st.session_state:
    h = load_history()
    st.session_state.file_history = {
        "현장명": h.get("현장명", []),
        "업체명": h.get("업체명", []),
        "위치":   h.get("위치",   []),
        "현장코드_map": h.get("현장코드_map", {}),
        "past_records": h.get("records", []),
    }

meta   = st.session_state.meta
fh     = st.session_state.file_history

if "past_active_recs" not in st.session_state:
    st.session_state.past_active_recs = []
if "past_sel_key" not in st.session_state:
    st.session_state.past_sel_key = None

# ── 사이드바 ───────────────────────────────────────────────────────────────
with st.sidebar:
    # 사이드바 전체 간격 압축
    st.markdown("""<style>
      section[data-testid="stSidebar"] .block-container { padding-top:0.8rem !important; }
      section[data-testid="stSidebar"] .stTextInput,
      section[data-testid="stSidebar"] .stSelectbox { margin-bottom:0 !important; }
      section[data-testid="stSidebar"] .stTextInput > div,
      section[data-testid="stSidebar"] .stSelectbox > div { margin-bottom:0.2rem !important; }
      section[data-testid="stSidebar"] hr { margin:0.4rem 0 !important; }
      section[data-testid="stSidebar"] h3 { margin:0.3rem 0 0.2rem !important; font-size:0.95rem !important; }
      section[data-testid="stSidebar"] p  { margin:0.15rem 0 !important; }
      .kosha-row { padding:2px 0 !important; }
    </style>""", unsafe_allow_html=True)

    st.markdown("### 📋 기본 정보")

    def meta_field(label, key, hist_key):
        opts = fh.get(hist_key, [])
        if opts:
            choice = st.selectbox(label, ["✏️ 직접 입력"] + opts, key=f"sel_{key}")
            if choice == "✏️ 직접 입력":
                val = st.text_input(" ", key=f"txt_{key}", label_visibility="collapsed",
                                    placeholder=label)
            else:
                val = choice
        else:
            val = st.text_input(label, key=f"txt_{key}", placeholder=label)
        return val

    # 현장코드 입력
    code_map = fh.get("현장코드_map", {})
    known_codes = list(code_map.keys())
    if known_codes:
        code_choice = st.selectbox("현장코드", ["✏️ 직접 입력"] + known_codes, key="sel_code")
        if code_choice == "✏️ 직접 입력":
            site_code = st.text_input(" ", key="txt_code", label_visibility="collapsed",
                                      placeholder="현장코드 (예: J881)")
        else:
            site_code = code_choice
    else:
        site_code = st.text_input("현장코드", key="txt_code", placeholder="현장코드 (예: J881)")

    # 현장코드 변경 시 현장명 자동입력
    prev_code = st.session_state.get("_prev_code", "")
    if site_code != prev_code:
        st.session_state["_prev_code"] = site_code
        if site_code in code_map:
            auto_name = code_map[site_code]
            site_opts = fh.get("현장명", [])
            if site_opts:
                if auto_name in site_opts:
                    st.session_state["sel_site"] = auto_name
                else:
                    st.session_state["sel_site"] = "✏️ 직접 입력"
                    st.session_state["txt_site"] = auto_name
            else:
                st.session_state["txt_site"] = auto_name
    meta["현장코드"] = site_code

    meta["현장명"] = meta_field("현장명", "site",    "현장명")
    meta["업체명"] = meta_field("업체명", "company", "업체명")
    meta["위치"]   = meta_field("온습도계 위치", "location", "위치")
    meta["측정자"] = st.text_input("측정자", key="txt_author", placeholder="홍길동",
                                    value=meta.get("측정자",""))

    st.markdown("---")
    st.markdown("### 📅 측정 기간")
    col_y, col_m = st.columns(2)
    with col_y:
        yr = st.selectbox("연도", list(range(2024, 2030)), index=list(range(2024,2030)).index(meta["year"]))
    with col_m:
        mo = st.selectbox("월", list(range(1,13)), index=meta["month"]-1,
                          format_func=lambda x: f"{x}월")
    meta["year"] = yr; meta["month"] = mo

    weeks = get_month_weeks(yr, mo)
    week_labels = [w[1] for w in weeks]
    cur_idx = next((i for i,w in enumerate(weeks) if w[0]==meta["week_n"]), 0)
    sel_wk = st.selectbox("주차", week_labels, index=cur_idx)
    chosen = weeks[week_labels.index(sel_wk)]
    meta["week_n"] = chosen[0]; meta["monday"] = chosen[2]
    monday: date = meta["monday"]

    if st.button("📂 이전기록 불러오기", use_container_width=True, type="primary"):
        _h = load_history()
        _mon_str = monday.isoformat()
        _week_dates = {(monday + timedelta(days=_d)).isoformat() for _d in range(7)}
        # 해당 주 기존 기록 제거 후 저장된 기록으로 교체
        st.session_state.records = [
            r for r in st.session_state.records
            if r.get("_date","") not in _week_dates
        ]
        for _r in _h.get("records", []):
            _rm = _r.get("_monday","")
            if not _rm and _r.get("_date"):
                try:
                    _d2 = date.fromisoformat(_r["_date"])
                    _rm = (_d2 - timedelta(days=_d2.weekday())).isoformat()
                except Exception:
                    pass
            if _rm == _mon_str:
                st.session_state.records.append(dict(_r,
                    _bytes=(base64.b64decode(_r["_bytes_b64"]) if _r.get("_bytes_b64") else None),
                    _done=True
                ))
        st.rerun()

    st.markdown("---")
    st.markdown("**📊 KOSHA 폭염 단계**")
    for dot, lbl, rng in [
        ("#34C759","관심","31~33°C"),
        ("#FFCC00","주의","33~35°C"),
        ("#FF9500","경고","35~38°C"),
        ("#FF3B30","위험","38°C 이상"),
    ]:
        st.markdown(
            f'<div class="kosha-row">'
            f'<div class="dot" style="background:{dot}"></div>'
            f'<b>{lbl}</b>&nbsp;<span style="color:#86868b">{rng}</span>'
            f'</div>', unsafe_allow_html=True)

    st.markdown("---")
    api_key = (st.secrets.get("GEMINI_API_KEY","") or os.environ.get("GEMINI_API_KEY",""))
    if api_key:
        st.success("✅ Gemini API 연결됨")
    else:
        st.error("❌ API 키 미설정")
        st.caption("`.streamlit/secrets.toml` 에 GEMINI_API_KEY를 입력하세요")

# ── 히어로 헤더 ────────────────────────────────────────────────────────────
week_start = monday; week_end = monday + timedelta(days=6)
wk_str = (f"{monday.year}년 {monday.month}월 {week_label_ko(meta['week_n'])}  "
          f"({monday.month}/{monday.day}~{week_end.month}/{week_end.day})")
_chips = []
if meta["현장명"]: _chips.append(f'<span class="chip">{meta["현장명"]}</span>')
if meta["업체명"]: _chips.append(f'<span class="chip">{meta["업체명"]}</span>')
if meta["위치"]:   _chips.append(f'<span class="chip">📍 {meta["위치"]}</span>')
_chips.append(f'<span class="chip">📅 {wk_str}</span>')
_chips_html = " ".join(_chips)

st.markdown(f"""
<div class="hero">
  <h1>🌡 체감온도 기록관리 대장</h1>
  <div class="hero-meta">{_chips_html}</div>
</div>
""", unsafe_allow_html=True)

def slot_from_filename(filename: str) -> str:
    m = re.search(r'\((\d+)\)', filename)
    if m:
        return {1:"오전1", 2:"오전2", 3:"오후1", 4:"오후2"}.get(int(m.group(1)), SLOTS[0])
    return SLOTS[0]

def _nan_to_none(v):
    return None if isinstance(v, float) and math.isnan(v) else v

DISPLAY_COLS = ["날짜","구분","측정시각","온도(°C)","습도(%)","체감온도(°C)","단계","조치사항","기타내용","측정자","비고"]

# ══════════════════════════════════════════════════════════ 기록 작성 ════════
uploaded = st.file_uploader(
    "📂  온습도계 사진 선택 (여러 장 동시 선택 가능, 업로드 시 자동 리사이즈)",
    type=["jpg","jpeg","png","bmp"], accept_multiple_files=True,
)

if uploaded:
    existing_names = {r["_filename"] for r in st.session_state.records}
    existing_slots = {(r["_date"], r["_slot"]): i for i, r in enumerate(st.session_state.records)}
    added = 0
    for f in uploaded:
        if f.name in existing_names:
            continue
        raw    = f.read()
        rbytes = resize_image(raw)
        d_str  = parse_date(f.name)
        slot   = slot_from_filename(f.name)
        d_obj  = date.fromisoformat(d_str) if d_str else monday
        rec_monday = d_obj - timedelta(days=d_obj.weekday())
        d_key  = d_obj.isoformat()
        sk = (d_key, slot)
        if sk in existing_slots:
            idx = existing_slots[sk]
            st.session_state.records[idx].update({"_bytes": rbytes, "_filename": f.name, "_done": False})
        else:
            st.session_state.records.append({
                "_filename": f.name, "_bytes": rbytes,
                "_date": d_key, "_slot": slot, "_done": False,
                "_monday": rec_monday.isoformat(),
                "날짜": f"{d_obj.month}/{d_obj.day}({DAYS_KO[d_obj.weekday()]})",
                "구분": slot, "측정시각": "", "온도(°C)": None,
                "습도(%)": None, "체감온도(°C)": None, "단계": "",
                "조치사항": "N/A", "기타내용": "",
                "측정자": meta.get("측정자",""),
                "비고": "",
            })
        added += 1
    if added:
        st.toast(f"{added}개 사진 추가 (리사이즈 완료)", icon="📸")

# ── 버튼 행 ──────────────────────────────────────────────────────────────
b1, b2, b3, b4, _ = st.columns([1.3, 1, 1, 1, 0.8])
with b1:
    do_extract = st.button("🤖  AI 자동 추출", type="primary",
                            disabled=(not api_key or not st.session_state.records),
                            use_container_width=True)
with b2:
    if st.button("🔄  미완료 재추출", use_container_width=True,
                 disabled=not st.session_state.records):
        for r in st.session_state.records: r["_done"] = False
        st.rerun()
with b3:
    if st.button("💾  최종 저장", use_container_width=True,
                 disabled=not any(r.get("_done") for r in st.session_state.records)):
        from collections import defaultdict as _ddict
        _wk_groups = _ddict(list)
        for _r in st.session_state.records:
            _wm = _r.get("_monday","")
            if not _wm and _r.get("_date"):
                try:
                    _d3 = date.fromisoformat(_r["_date"])
                    _wm = (_d3 - timedelta(days=_d3.weekday())).isoformat()
                except Exception:
                    _wm = monday.isoformat()
            _wk_groups[_wm].append(_r)
        _saved = 0
        for _wm, _recs in _wk_groups.items():
            if any(_r.get("_done") for _r in _recs):
                try: _wmon = date.fromisoformat(_wm)
                except Exception: _wmon = monday
                save_history(meta, _recs, _wmon)
                _saved += 1
        _h2 = load_history()
        st.session_state.file_history["past_records"] = _h2.get("records", [])
        st.session_state.file_history["현장코드_map"] = _h2.get("현장코드_map", {})
        st.toast(f"✅ {_saved}개 주차 저장 완료!", icon="💾")
with b4:
    if st.button("🗑  전체 초기화", use_container_width=True):
        st.session_state.records = []; st.rerun()

# ── AI 추출 ──────────────────────────────────────────────────────────────
if do_extract and api_key:
    pending = [r for r in st.session_state.records if not r["_done"] and r.get("_bytes")]
    if not pending:
        st.info("추출할 항목이 없습니다.")
    else:
        client = google_genai.Client(api_key=api_key)
        bar  = st.progress(0); stat = st.empty()
        stat.caption(f"⏳ {len(pending)}장 병렬 추출 중…")
        errors = []; done_count = 0
        with ThreadPoolExecutor(max_workers=min(4, len(pending))) as ex:
            futures = {ex.submit(_extract_one, client, rec): rec for rec in pending}
            for fut in as_completed(futures):
                rec = futures[fut]; done_count += 1
                bar.progress(done_count / len(pending))
                try:
                    fut.result()
                    pct = int(done_count / len(pending) * 100)
                    stat.caption(f"⏳ {done_count}/{len(pending)} ({pct}%) – {rec['_filename']}")
                except Exception as e:
                    errors.append(rec["_filename"])
                    st.error(f"❌ **{rec['_filename']}** 오류: {e}")
        bar.progress(1.0)
        stat.success("✅ 추출 완료!" if not errors else f"⚠️ {len(errors)}개 실패")
        st.rerun()

# ── 주별 그룹화 헬퍼 ────────────────────────────────────────────────────
def _rec_monday_wk(r):
    ds = r.get("_monday") or r.get("_date", "")
    if ds:
        try:
            d = date.fromisoformat(ds[:10])
            return d - timedelta(days=d.weekday())
        except:
            pass
    return monday

all_mondays = sorted({_rec_monday_wk(r) for r in st.session_state.records}) if st.session_state.records else [monday]

# ── KPI 요약 (전체) ──────────────────────────────────────────────────────
done_recs = [r for r in st.session_state.records if r["_done"]]
if done_recs:
    temps = [r["온도(°C)"]    for r in done_recs if r["온도(°C)"]    is not None]
    feels = [r["체감온도(°C)"] for r in done_recs if r["체감온도(°C)"] is not None]
    humid = [r["습도(%)"]     for r in done_recs if r["습도(%)"]     is not None]
    c1,c2,c3,c4 = st.columns(4)
    for col, val, lbl, cls in [
        (c1, len(done_recs), "측정 건수", "kpi-card-soft"),
        (c2, f"{max(temps):.1f}°C" if temps else "-", "최고 온도", "kpi-card"),
        (c3, f"{max(humid):.0f}%"  if humid else "-", "최고 습도", "kpi-card"),
        (c4, f"{max(feels):.1f}°C" if feels else "-", "최고 체감온도", "kpi-card-hl"),
    ]:
        col.markdown(f'<div class="{cls}"><div class="kpi-val">{val}</div>'
                     f'<div class="kpi-lbl">{lbl}</div></div>', unsafe_allow_html=True)
    st.markdown("")

# ── 주별 섹션 ────────────────────────────────────────────────────────────
changed = False
for wk_monday in all_mondays:
    wk_end  = wk_monday + timedelta(days=6)
    wk_info = get_month_weeks(wk_monday.year, wk_monday.month)
    wk_n    = next((w[0] for w in wk_info if w[2] == wk_monday), 1)
    this_wk_str = (f"{wk_monday.year}년 {wk_monday.month}월 {week_label_ko(wk_n)}  "
                   f"({wk_monday.month}/{wk_monday.day}~{wk_end.month}/{wk_end.day})")

    wk_done = [r for r in st.session_state.records
               if _rec_monday_wk(r) == wk_monday and r["_done"]]

    hc1, hc2 = st.columns([4, 1])
    with hc1:
        st.markdown(f'<div class="week-tag">📅 {this_wk_str}</div>', unsafe_allow_html=True)
    with hc2:
        if wk_done:
            wk_all = [r for r in st.session_state.records if _rec_monday_wk(r) == wk_monday]
            _code  = meta.get("현장코드","").strip()
            _loc   = meta.get("위치","").strip() or "위치미정"
            _cpfx  = f"({_code})" if _code else ""
            fname  = f"{_cpfx}체감온도 기록관리 대장_{_loc}_{wk_monday.year}년{wk_monday.month}월{week_label_ko(wk_n)}.xlsx"
            save_history(meta, wk_all, wk_monday)
            st.download_button("💾 엑셀", key=f"dl_{wk_monday.isoformat()}",
                data=build_excel(wk_all, meta, wk_monday),
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        else:
            st.button("💾 엑셀", disabled=True, key=f"dl_dis_{wk_monday.isoformat()}",
                      use_container_width=True)

    # 주간 차트
    wk_hi_map = {(r.get("_date",""), r.get("_slot","")): r.get("체감온도(°C)")
                 for r in wk_done}
    if wk_hi_map:
        day_labels = [
            f"{(wk_monday+timedelta(days=di)).month}/{(wk_monday+timedelta(days=di)).day}"
            f"({DAYS_KO[(wk_monday+timedelta(days=di)).weekday()]})"
            for di in range(7)
        ]
        slot_colors = {"오전1":"#4A90E2","오전2":"#34C759","오후1":"#FF9500","오후2":"#FF3B30"}
        fig = go.Figure()
        for slot in SLOTS:
            y_vals = [wk_hi_map.get(((wk_monday+timedelta(days=di)).isoformat(), slot))
                      for di in range(7)]
            fig.add_trace(go.Bar(
                name=slot, x=day_labels, y=y_vals,
                marker_color=slot_colors[slot], opacity=0.85,
                text=[f"{v:.1f}" if v is not None else "" for v in y_vals],
                textposition="outside", textfont=dict(size=9),
            ))
        for lbl, col, val in [("관심","#34C759",31),("주의","#FFCC00",33),
                               ("경고","#FF9500",35),("위험","#FF3B30",38)]:
            fig.add_hline(y=val, line_dash="dot", line_color=col, line_width=1.5,
                          annotation_text=f" {lbl} {val}°C", annotation_position="right",
                          annotation_font=dict(size=9, color=col))
        all_vals = [v for v in wk_hi_map.values() if v is not None]
        y_min = min(min(all_vals) - 3, 28) if all_vals else 25
        y_max = max(max(all_vals) + 4, 42) if all_vals else 42
        fig.update_layout(
            barmode="group", height=260,
            margin=dict(l=0, r=80, t=8, b=0),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            yaxis=dict(title="체감온도(°C)", range=[y_min, y_max], gridcolor="#f0f0f0"),
            xaxis=dict(tickfont=dict(size=11)),
            plot_bgcolor="white", paper_bgcolor="white",
        )
        st.markdown('<div class="mac-card">', unsafe_allow_html=True)
        st.plotly_chart(fig, use_container_width=True, key=f"chart_{wk_monday.isoformat()}")
        st.markdown('</div>', unsafe_allow_html=True)

    # 주간 데이터 그리드
    rec_map_wk: dict[tuple, int] = {}
    for idx, r in enumerate(st.session_state.records):
        if _rec_monday_wk(r) == wk_monday:
            rec_map_wk[(r.get("_date",""), r.get("_slot",""))] = idx

    grid_rows = []
    for di in range(7):
        d = wk_monday + timedelta(days=di)
        for slot in SLOTS:
            key = (d.isoformat(), slot)
            if key in rec_map_wk:
                r = st.session_state.records[rec_map_wk[key]]
                grid_rows.append({
                    "날짜": f"{d.month}/{d.day}({DAYS_KO[d.weekday()]})",
                    "구분": slot,
                    "측정시각": r.get("측정시각",""),
                    "온도(°C)": r.get("온도(°C)"),
                    "습도(%)": r.get("습도(%)"),
                    "체감온도(°C)": r.get("체감온도(°C)"),
                    "단계": LEVEL_BADGE.get(r.get("단계",""), r.get("단계","")),
                    "조치사항": r.get("조치사항","N/A"),
                    "기타내용": r.get("기타내용",""),
                    "측정자": r.get("측정자",""),
                    "비고": r.get("비고",""),
                    "_key": key,
                })
            else:
                grid_rows.append({
                    "날짜": f"{d.month}/{d.day}({DAYS_KO[d.weekday()]})",
                    "구분": slot,
                    "측정시각": "", "온도(°C)": None, "습도(%)": None,
                    "체감온도(°C)": None, "단계": "", "조치사항": "N/A",
                    "기타내용": "", "측정자": meta.get("측정자",""), "비고": "",
                    "_key": key,
                })

    grid_keys = [r["_key"] for r in grid_rows]
    df = pd.DataFrame([{c: r[c] for c in DISPLAY_COLS} for r in grid_rows])
    df.insert(0, "삭제", False)

    edited = st.data_editor(
        df, use_container_width=True, num_rows="fixed", hide_index=True,
        key=f"editor_{wk_monday.isoformat()}",
        column_config={
            "삭제":         st.column_config.CheckboxColumn("🗑", default=False, width=40),
            "날짜":         st.column_config.TextColumn("날짜",     disabled=True, width=90),
            "구분":         st.column_config.TextColumn("구분",     disabled=True, width=70),
            "측정시각":     st.column_config.TextColumn("측정시각", width=80),
            "온도(°C)":     st.column_config.NumberColumn("온도(°C)",    format="%.1f°C", step=0.1, width=85),
            "습도(%)":      st.column_config.NumberColumn("습도(%)",     format="%.0f%%", step=1,   width=75),
            "체감온도(°C)": st.column_config.NumberColumn("체감온도(°C)", format="%.1f°C", disabled=True, width=100),
            "단계":         st.column_config.TextColumn("단계",     disabled=True, width=60),
            "조치사항":     st.column_config.SelectboxColumn("조치사항", options=ACTIONS, width=150),
            "기타내용":     st.column_config.TextColumn("기타내용(기타 선택시)", width=130),
            "측정자":       st.column_config.TextColumn("측정자",   width=80),
            "비고":         st.column_config.TextColumn("비고",     width=130),
        },
    )

    # 체크된 행 삭제
    del_indices = [i for i in range(len(grid_keys)) if edited.iloc[i]["삭제"]]
    if del_indices:
        del_keys = {grid_keys[i] for i in del_indices}
        if st.button(f"🗑 선택한 {len(del_indices)}행 초기화",
                     key=f"del_{wk_monday.isoformat()}", type="secondary"):
            st.session_state.records = [
                r for r in st.session_state.records
                if (r.get("_date",""), r.get("_slot","")) not in del_keys
            ]
            st.rerun()

    for row_i, key in enumerate(grid_keys):
        row = {k: _nan_to_none(v) for k, v in edited.iloc[row_i].to_dict().items()}
        T  = row["온도(°C)"]; RH = row["습도(%)"]
        if T is not None and RH is not None:
            try:
                hi = heat_index(float(T), float(RH))
                row["체감온도(°C)"] = hi
                row["단계"] = heat_label(hi)
            except Exception:
                pass

        if key in rec_map_wk:
            r = st.session_state.records[rec_map_wk[key]]
            for c in DISPLAY_COLS:
                old_val = _nan_to_none(r.get(c))
                if c not in ("날짜","구분","체감온도(°C)","단계") and old_val != row.get(c):
                    r[c] = row.get(c); changed = True
            r["체감온도(°C)"] = row.get("체감온도(°C)")
            r["단계"] = heat_label(row["체감온도(°C)"]) if row.get("체감온도(°C)") is not None else ""
        else:
            if T is not None and RH is not None:
                d_str, slot = key
                try: d_obj = date.fromisoformat(d_str)
                except: d_obj = wk_monday
                st.session_state.records.append({
                    "_filename": "", "_bytes": None,
                    "_date": d_str, "_slot": slot, "_done": True,
                    "_monday": wk_monday.isoformat(),
                    "날짜": f"{d_obj.month}/{d_obj.day}({DAYS_KO[d_obj.weekday()]})",
                    "구분": slot, **{c: row[c] for c in DISPLAY_COLS if c not in ("날짜","구분")},
                })
                changed = True

if changed:
    st.rerun()

# ── 사진 미리보기 ────────────────────────────────────────────────────────
photo_recs = [r for r in st.session_state.records if r.get("_bytes")]
if photo_recs:
    with st.expander("📷 사진 미리보기", expanded=False):
        cols = st.columns(min(4, len(photo_recs)))
        for i, rec in enumerate(photo_recs):
            with cols[i % 4]:
                try:
                    img = Image.open(io.BytesIO(rec["_bytes"]))
                    lev = rec.get("단계","")
                    cap = f"**{rec['_filename']}**\n{rec.get('구분','')}"
                    if rec.get("온도(°C)") is not None:
                        cap += f"\n{rec['온도(°C)']}°C / {rec['습도(%)']}%"
                    if rec.get("체감온도(°C)") is not None:
                        cap += f" → 체감 **{rec['체감온도(°C)']}°C** ({lev})"
                    st.image(img, use_container_width=True)
                    st.caption(cap)
                except Exception:
                    st.caption(rec.get("_filename",""))

